import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def calculate_improvement(baseline_val, comparison_val, metric):
    """
    Calculate percentage improvement from baseline to comparison.
    
    For RMSE, MAE, MAPE, AIC, BIC: Lower is better
    For R2: Higher is better
    
    Returns improvement % (positive = good, negative = bad)
    """
    if pd.isna(baseline_val) or pd.isna(comparison_val):
        return np.nan
    
    if baseline_val == 0:
        return np.nan
    
    if metric == 'R2':
        # Higher is better
        improvement = ((comparison_val - baseline_val) / abs(baseline_val)) * 100
    else:
        # Lower is better (RMSE, MAE, MAPE, AIC, BIC)
        improvement = ((baseline_val - comparison_val) / baseline_val) * 100
    
    return improvement


def create_comparison_table(results_df, metric, baseline_label, exog_label, eemd_label):
    """
    Create comparison table for a single metric.
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Combined results from all experiments
    metric : str
        Metric name (RMSE, MAE, MAPE, R2, AIC, BIC)
    baseline_label, exog_label, eemd_label : str
        Labels used in the Model column (e.g., 'Sales-Baseline')
    
    Returns
    -------
    comparison_df : pd.DataFrame
        Formatted comparison table
    """
    # Extract model names (remove the label suffix)
    results_df = results_df.copy()
    results_df['ModelName'] = results_df['Model'].str.replace(f'-{baseline_label}', '', regex=False)
    results_df['ModelName'] = results_df['ModelName'].str.replace(f'-{exog_label}', '', regex=False)
    results_df['ModelName'] = results_df['ModelName'].str.replace(f'-{eemd_label}', '', regex=False)
    
    # Separate results by scenario
    baseline_df = results_df[results_df['Model'].str.contains(baseline_label)].copy()
    exog_df = results_df[results_df['Model'].str.contains(exog_label)].copy()
    eemd_df = results_df[results_df['Model'].str.contains(eemd_label)].copy()
    
    # Get unique model names
    model_names = baseline_df['ModelName'].unique()
    
    # Build comparison table
    comparison_data = []
    
    for model in model_names:
        baseline_val = baseline_df[baseline_df['ModelName'] == model][metric].values
        exog_val = exog_df[exog_df['ModelName'] == model][metric].values
        eemd_val = eemd_df[eemd_df['ModelName'] == model][metric].values
        
        baseline_val = baseline_val[0] if len(baseline_val) > 0 else np.nan
        exog_val = exog_val[0] if len(exog_val) > 0 else np.nan
        eemd_val = eemd_val[0] if len(eemd_val) > 0 else np.nan
        
        # Calculate improvements
        exog_imp = calculate_improvement(baseline_val, exog_val, metric)
        eemd_imp = calculate_improvement(baseline_val, eemd_val, metric)
        
        # Format values with improvement percentages
        if not pd.isna(exog_val):
            exog_str = f"{exog_val:.3f} ({exog_imp:+.1f}%)" if not pd.isna(exog_imp) else f"{exog_val:.3f}"
        else:
            exog_str = "N/A"
        
        if not pd.isna(eemd_val):
            eemd_str = f"{eemd_val:.3f} ({eemd_imp:+.1f}%)" if not pd.isna(eemd_imp) else f"{eemd_val:.3f}"
        else:
            eemd_str = "N/A"
        
        comparison_data.append({
            'Model': model,
            'Baseline': f"{baseline_val:.3f}" if not pd.isna(baseline_val) else "N/A",
            'Exogenous': exog_str,
            'EEMD': eemd_str,
            # FIXED: Remove underscore prefix (pandas namedtuples don't support it)
            'exog_imp': exog_imp,
            'eemd_imp': eemd_imp
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    return comparison_df


def create_styled_excel(results_df, output_path, baseline_label, exog_label, eemd_label):
    """
    Create Excel file with comparison tables for all metrics with proper formatting.
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Combined results DataFrame
    output_path : str
        Path to save Excel file
    baseline_label, exog_label, eemd_label : str
        Scenario labels
    """
    metrics = ['RMSE', 'MAE', 'MAPE', 'R2', 'AIC', 'BIC']
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    improvement_good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    improvement_bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for metric in metrics:
        # Create comparison table
        comparison_df = create_comparison_table(results_df, metric, baseline_label, exog_label, eemd_label)
        
        # Create worksheet
        ws = wb.create_sheet(title=metric)
        
        # Add title
        ws['A1'] = f"{metric} Comparison: Baseline vs Exogenous vs EEMD"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Add headers
        headers = ['Model', 'Baseline', 'Exogenous (Δ%)', 'EEMD (Δ%)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Add data
        for row_idx, row_data in enumerate(comparison_df.itertuples(index=False), start=4):
            model = row_data.Model
            baseline = row_data.Baseline
            exog = row_data.Exogenous
            eemd = row_data.EEMD
            exog_imp = row_data.exog_imp  # FIXED: removed underscore
            eemd_imp = row_data.eemd_imp  # FIXED: removed underscore
            
            # Model name
            cell = ws.cell(row=row_idx, column=1, value=model)
            cell.border = border_thin
            cell.font = Font(bold=True)
            
            # Baseline value
            cell = ws.cell(row=row_idx, column=2, value=baseline)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
            
            # Exogenous value with improvement
            cell = ws.cell(row=row_idx, column=3, value=exog)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
            
            # Color code based on improvement
            if not pd.isna(exog_imp):
                if exog_imp > 0:
                    cell.fill = improvement_good_fill
                    cell.font = Font(color="006100")
                elif exog_imp < 0:
                    cell.fill = improvement_bad_fill
                    cell.font = Font(color="9C0006")
            
            # EEMD value with improvement
            cell = ws.cell(row=row_idx, column=4, value=eemd)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
            
            # Color code based on improvement
            if not pd.isna(eemd_imp):
                if eemd_imp > 0:
                    cell.fill = improvement_good_fill
                    cell.font = Font(color="006100")
                elif eemd_imp < 0:
                    cell.fill = improvement_bad_fill
                    cell.font = Font(color="9C0006")
        
        # Add summary statistics at bottom
        summary_row = len(comparison_df) + 5
        
        ws.cell(row=summary_row, column=1, value="Average Improvement:").font = Font(bold=True)
        
        # Calculate average improvements - FIXED: removed underscores
        avg_exog = comparison_df['exog_imp'].mean()
        avg_eemd = comparison_df['eemd_imp'].mean()
        
        ws.cell(row=summary_row, column=2, value="Baseline")
        ws.cell(row=summary_row, column=3, value=f"{avg_exog:+.2f}%")
        ws.cell(row=summary_row, column=4, value=f"{avg_eemd:+.2f}%")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary['A1'] = "Model Performance Summary"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    summary_data = []
    for metric in metrics:
        comparison_df = create_comparison_table(results_df, metric, baseline_label, exog_label, eemd_label)
        # FIXED: removed underscores
        avg_exog = comparison_df['exog_imp'].mean()
        avg_eemd = comparison_df['eemd_imp'].mean()
        
        summary_data.append({
            'Metric': metric,
            'Avg Exog Improvement (%)': f"{avg_exog:+.2f}",
            'Avg EEMD Improvement (%)': f"{avg_eemd:+.2f}"
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    for row_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
    
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 25
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Comparison tables saved to: {output_path}")
